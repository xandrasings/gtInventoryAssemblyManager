from .const import *
from .fileManagement import *
from .sys import *
from ..Classes.Component import *
from ..Classes.Task import *


from openpyxl import load_workbook

import os

def loadWorkbook(fileType, filePath, readOnly = False):
	workBook = load_workbook(filePath, readOnly)
	assertExpectedWorkSheets(fileType, workBook)
	return workBook


def assertExpectedWorkSheets(fileType, workBook):
	for sheet in FILE_SHEETS[fileType]:
		if not sheet in workBook.sheetnames:
			print('Selected task file is missing ' + sheet + ' worksheet.')
			quit()


def generateTaskList():
	taskList = []

	workBook = loadWorkbook(TASKS, selectTaskFile(), True)
	taskSheet = workBook[TASK_FILE_TASK_SHEET]

	productDictionary = generateProductDictionary(workBook)

	for i in range(2, taskSheet.max_row):
		product = str(taskSheet.cell(row = i, column = 1).value)
		if product == "None":
			break
		quantity = taskSheet.cell(row = i, column = 2).value
		components = productDictionary[product]
		taskList.append(Task(product, quantity, components))

	return taskList


def generateProductDictionary(workBook):
	componentSheet = workBook[TASK_FILE_COMPONENTS_SHEET]
	# TODO assert worksheet dimensions from ws.calculate_dimension() sheet.get_highest_column()
	# TODO assert column headers are as expected
	# TODO use regex to assert expected formatting of rows
	productDictionary = {}
	for i in range(2, componentSheet.max_row):
		product = str(componentSheet.cell(row = i, column = 1).value)
		part = str(componentSheet.cell(row = i, column = 2).value)
		quantity = componentSheet.cell(row = i, column = 3).value

		if product not in productDictionary.keys():
			productDictionary[product] = []

		productDictionary[product].append(Component(part, quantity))

	return productDictionary


def generateAssemblyOrderFile(taskList):
	createAssemblyOrderFile()
	try:
		populateAssemblyOrderFile(taskList)
		renameAssemblyOrderFile()
	except:
		print('Something went wrong! Cancelling process.')
		deleteAssemblyOrderFile()
		quit()


def populateAssemblyOrderFile(taskList):
	workBook = loadWorkbook(TEMPLATE, ASSEMBLY_ORDER_FILE_NAME_PATH)
	templateSheet = workBook[TEMPLATE_FILE_TEMPLATE_SHEET]
	taskCount = len(taskList)

	for i in range (0, taskCount):
		product = taskList[i].getProduct()
		newSheet = workBook.copy_worksheet(templateSheet)
		newSheet.title = generateSheetName(i, product)
		#TODO nerd out
		newSheet.cell(row = 1, column = 7).value = generateFraction(i, taskCount);
		newSheet.cell(row = 2, column = 4).value = product
		newSheet.cell(row = 4, column = 4).value = taskList[i].getQuantity()

		#TODO implement max
		currentRow = 8
		for component in taskList[i].getComponents():
			newSheet.cell(row = currentRow, column = 1).value = component.getPart()
			newSheet.cell(row = currentRow, column = 4).value = component.getQuantity()
			currentRow = currentRow + 1


	workBook.remove(templateSheet)
	workBook.save(ASSEMBLY_ORDER_FILE_NAME_PATH)


def generateSheetName(counter, product):
	return str(counter + 1) + " " + product[:27]


def generateFraction(counter, total):
	return str(counter + 1) + "/" + str(total)